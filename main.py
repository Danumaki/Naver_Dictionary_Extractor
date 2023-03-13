from re import findall
from regex import search
from pdfminer.high_level import extract_pages
from pdfminer.layout import LTTextContainer, LTTextLine, LTChar
from tkinter import filedialog
# from pprint import pprint as pp
import json
import openpyxl  # from openpyxl import Workbook, load_workbook
# from openpyxl.styles import Alignment
import os
from glob import glob


def contains_hangeul(text):
    """Returns True if char is hangeul. If for some reason char is a longer string, then return True if there is
    any hangeul in the string"""
    if search(r'\p{IsHangul}', text):
        return True
    return False


def is_all_hangeul(text):
    return text.strip() == "".join([x for x in text if contains_hangeul(x)]).strip()


def contains_chinese(text):
    """Returns True, if txt contains a Chinese character"""
    if findall(r"[\u4e00-\u9fff]+", text):
        return True
    return False


def get_dirname():
    """
    Props the user to select a folder, and we capture its path in a string

        Returns:
                file_location (str): Path of selected file
    """
    dir_location = filedialog.askdirectory(
        initialdir="E:\\Programming_HQ\\Python\\Projects\\Naver_Dictionary_Extractor",
        title="Choose directory of .pdf workbooks.")
    return dir_location


def read_pdf_page(filename):
    all_doc = []

    for page_layout in extract_pages(filename):
        for element in page_layout:
            if isinstance(element, LTTextContainer):
                for text_line in element:
                    if isinstance(text_line, LTTextLine):
                        for char in text_line:
                            if isinstance(char, LTChar):
                                char_size = char.size
                all_doc.append([char_size, (element.get_text())])

    return all_doc


def write_to_excel(word_list, filename):
    pdf_name_titled = filename.title()

    try:
        workbook = openpyxl.load_workbook("Naver Korean-English Workbooks.xlsx")
        if pdf_name_titled in workbook.sheetnames:
            pass  # append
        else:
            workbook.create_sheet(pdf_name_titled)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = pdf_name_titled
    finally:
        worksheet = workbook[pdf_name_titled]

    worksheet.cell(row=1, column=1).value = "Angol"
    worksheet.cell(row=1, column=2).value = "Koreai"
    worksheet.cell(row=1, column=3).value = "tudás_angol"
    worksheet.cell(row=1, column=4).value = "tudás_koreai"
    worksheet.cell(row=1, column=5).value = "átírás + memo"
    worksheet.cell(row=1, column=6).value = "description"

    line_written = 1
    for word in word_list:
        single_memo = "[" + word["romanization"] + "]"
        if word["memo"]:
            single_memo += "\nMemo: " + word["memo"]
        meaning_no = 0
        for meaning in word["meanings"]:
            if meaning["meaning"]:
                meaning_no += 1
                worksheet.cell(row=line_written + 1, column=1).value = meaning["meaning"]
                worksheet.cell(row=line_written + 1, column=2).value = word["Korean"]
                worksheet.cell(row=line_written + 1, column=5).value = single_memo
                if meaning["description_English"]:
                    worksheet.cell(row=line_written + 1, column=6).value = meaning["description_English"].strip()
                # cell = "E{}".format(line_written + 1)
                # worksheet[cell].alignment = Alignment(wrap_text=True)
                line_written += 1

    workbook.save(os.path.join(selected_path, "Naver Korean-English Workbooks.xlsx"))
    workbook.close()


def word_extractor(full_document):
    upcoming_memo = False
    words = json.load(open(project_path + "\\words.json"))

    for element in full_document:
        # print(element)
        if 9.8 < element[0]:
            if element[1].split()[0][-1].isnumeric():
                global_meaning_no = element[1].split()[0][-1]
                element[1] = element[1].replace(global_meaning_no, "", 1)
            if "[" in element[1]:
                kor_cluster = element[1].strip("]").split("[")
                if not kor_cluster[0]:  # if it starts with "[" (like: [kakteil]), then add to previous element
                    words[-2]["romanization"] = kor_cluster[1]
                elif not contains_chinese(kor_cluster[0]):
                    words[-1]["Korean"] = \
                        "".join([x for x in kor_cluster[0].strip().split() if is_all_hangeul(x)]).strip()
                    words[-1]["romanization"] = kor_cluster[1].strip("]\n")
                    words.extend(json.load(open(project_path + "\\words.json")))
                else:
                    words[-1]["Korean"] = \
                        "".join([x for x in kor_cluster[0].strip().split() if is_all_hangeul(x)]).strip()
                    words[-1]["Chinese_characterization"] = \
                        "".join([x for x in kor_cluster[0].strip().split() if contains_chinese(x)]).strip()
                    words[-1]["romanization"] = kor_cluster[1].strip("]\n")
                    words.extend(json.load(open(project_path + "\\words.json")))

            else:
                kor_cluster = element[1].strip().split()
                if not contains_chinese(element[1]):
                    words[-1]["Korean"] = \
                        "".join([x for x in kor_cluster if is_all_hangeul(x)]).strip()
                    words.extend(json.load(open(project_path + "\\words.json")))
                else:
                    words[-1]["Korean"] = \
                        "".join([x for x in kor_cluster[0].strip().split() if is_all_hangeul(x)]).strip()
                    words[-1]["Chinese_characterization"] = \
                        "".join([x for x in kor_cluster[0].strip().split() if contains_chinese(x)]).strip()
                    words.extend(json.load(open(project_path + "\\words.json")))

        elif 8.4 < element[0] < 8.5:
            if not upcoming_memo:
                meaning_no = 0
                if element[1].split()[0].strip(".").isnumeric():
                    meaning_no = int(element[1].split()[0].strip("."))
                if meaning_no and element[1].split()[0][-1] == ".":
                    meaning_cluster = element[1].strip().split()
                    if meaning_no > 1:
                        words[-2]["meanings"].extend(json.load(open(project_path + "\\words.json"))[0]["meanings"])
                    if element[1].strip("\n").count("\n") >= 1:
                        everything_cluster = element[1].strip("\n").split("\n")
                        # Ha van megadva szófaj (pl. [명사]), de nem az angol jelentésben van a []:
                        if "[" in element[1] and is_all_hangeul(element[1][element[1].index("[") + 1]):
                            words[-2]["meanings"][-1]["type"] = \
                                everything_cluster[0].split()[1][1:-1]
                            if not contains_hangeul("".join(everything_cluster[0].strip().split("]")[1]).strip()):
                                words[-2]["meanings"][-1]["meaning"] = \
                                    "".join(everything_cluster[0].strip().split("]")[1]).strip()
                            words[-2]["meanings"][-1]["description_Korean"] = \
                                "".join(everything_cluster[1].strip()).strip()
                            if len(everything_cluster) == 3:
                                words[-2]["meanings"][-1]["description_English"] = \
                                    "".join(everything_cluster[2].strip()).strip()
                        else:
                            if not contains_hangeul("".join(everything_cluster[0].strip()).strip()):
                                words[-2]["meanings"][-1]["meaning"] = \
                                    "".join(everything_cluster[0].strip()).strip()
                            words[-2]["meanings"][-1]["description_Korean"] = \
                                "".join(everything_cluster[1].strip()).strip()
                            if len(everything_cluster) == 3:
                                words[-2]["meanings"][-1]["description_English"] = \
                                    "".join(everything_cluster[2].strip()).strip()
                    else:
                        # Ha van megadva szófaj (pl. [명사]), de nem az angol jelentésben van a []:
                        if "[" in element[1] and is_all_hangeul(element[1][element[1].index("[") + 1]):
                            words[-2]["meanings"][-1]["type"] = \
                                meaning_cluster[1][1:-1]
                            if not contains_hangeul(" ".join(meaning_cluster[2:]).strip()):
                                words[-2]["meanings"][-1]["meaning"] = \
                                    " ".join(meaning_cluster[2:]).strip()
                        else:
                            if not contains_hangeul(" ".join(meaning_cluster[1:]).strip()):
                                words[-2]["meanings"][-1]["meaning"] = \
                                    " ".join(meaning_cluster[1:]).strip()
                else:
                    description_cluster = element[1].strip().split("\n")
                    for des in description_cluster:
                        if contains_hangeul(des):
                            words[-2]["meanings"][-1]["description_Korean"] = des.strip()
                        else:
                            words[-2]["meanings"][-1]["description_English"] = des.strip()
            else:
                words[-2]["memo"] = element[1].strip()
                upcoming_memo = False

        elif 7.0 < element[0] < 7.1:
            if element[1] != "Memo\n":
                if words[-2]["meanings"][-1]["examples"][-1]:
                    words[-2]["meanings"][-1]["examples"].extend \
                        (json.load(open(project_path + "\\words.json"))[0]["meanings"][0]["examples"])
                words[-2]["meanings"][-1]["examples"][-1] = element[1].strip()
            else:
                upcoming_memo = True

    del words[-1]
    # pp(words)

    return words


def run_on_all_files_in_dir():
    global selected_path

    selected_path = get_dirname()
    print(selected_path)
    os.chdir(selected_path)

    all_pdf_files = glob("*.pdf")
    for file in all_pdf_files:
        print(file)
        write_to_excel(word_extractor(read_pdf_page(file)), file)


if __name__ == '__main__':
    project_path = os.path.dirname(__file__)
    selected_path = os.path.dirname(__file__)

    run_on_all_files_in_dir()
